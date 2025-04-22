import os
import json
import logging
import re  
from datetime import datetime, date, timezone, timedelta
from logging.handlers import RotatingFileHandler
from zoneinfo import ZoneInfo
from flask import (
    Flask, render_template, request, redirect, 
    session, send_file, send_from_directory, 
    flash, url_for, jsonify
)
from flask_sqlalchemy import SQLAlchemy
from openpyxl.utils import get_column_letter
import pandas as pd
import openpyxl
from apscheduler.schedulers.background import BackgroundScheduler
from io import BytesIO
from sqlalchemy.exc import SQLAlchemyError, IntegrityError
from sqlalchemy import func, text
from flask_migrate import Migrate

app = Flask(__name__, static_folder='static', template_folder='templates')

basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-production-secret-key')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'data.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB文件上传限制
app.config['UPLOAD_FOLDER'] = os.path.join(basedir, 'uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)
migrate = Migrate(app, db)

if not app.debug:
    if not os.path.exists('logs'):
        os.mkdir('logs')
    file_handler = RotatingFileHandler(
        'logs/app.log',
        maxBytes=1024*1024*10,
        backupCount=5
    )
    file_handler.setFormatter(logging.Formatter(
        '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
    ))
    file_handler.setLevel(logging.INFO)
    app.logger.addHandler(file_handler)
    app.logger.setLevel(logging.INFO)
    app.logger.info('Application startup')

# -------------------- 数据库模型 --------------------
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(20), unique=True, nullable=False)
    password = db.Column(db.String(60), nullable=False)
    department = db.Column(db.String(50))
    phone = db.Column(db.String(20))
    is_admin = db.Column(db.Boolean, default=False)
    is_department_info = db.Column(db.Boolean, default=False)
    is_department_head = db.Column(db.Boolean, default=False)  # 部门负责人标记
    is_company_info = db.Column(db.Boolean, default=False)
    is_general_dept_head = db.Column(db.Boolean, default=False)  # 集团部门负责人标记
    is_company_leader = db.Column(db.Boolean, default=False)

    @property
    def is_dept_head(self):
        """综合判断是否具备部门负责人权限"""
        return self.is_department_head or self.is_general_dept_head

class WeeklyReport(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    reporter = db.Column(db.String(20), nullable=False)
    department = db.Column(db.String(50), nullable=False)
    content = db.Column(db.Text)
    report_type = db.Column(db.String(20), nullable=False)
    status = db.Column(db.String(10))
    submit_time = db.Column(db.DateTime, default=datetime.utcnow)
    report_date = db.Column(db.Date)

class SupervisionProject(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_name = db.Column(db.String(100), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class SupervisionDetail(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('supervision_project.id'), nullable=False)
    main_content = db.Column(db.String(200), nullable=False)
    key_node = db.Column(db.String(100))
    responsible_dept = db.Column(db.String(50))
    responsible_person = db.Column(db.String(20))
    cooperating_dept = db.Column(db.String(200))
    cooperating_persons = db.Column(db.String(200))
    responsible_leader = db.Column(db.String(20))
    deadline = db.Column(db.Date)
    completion_time = db.Column(db.Date)
    status = db.Column(db.String(10), default='进行中')
    last_status_update = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None))
    project = db.relationship('SupervisionProject', backref='details')

class ProgressRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    detail_id = db.Column(db.Integer, db.ForeignKey('supervision_detail.id'), nullable=False)
    submitter = db.Column(db.String(20), nullable=False)
    content = db.Column(db.Text, nullable=False)
    submit_time = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None))
    status = db.Column(db.String(10), default='待审核')
    reviewer = db.Column(db.String(20))

class ProjectPlan(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    plan_name = db.Column(db.String(100), nullable=False)
    category = db.Column(db.String(20), default='项目类')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class TimePlan(db.Model):
    __tablename__ = 'time_plan'
    id = db.Column(db.Integer, primary_key=True)
    plan_name = db.Column(db.String(100), unique=True)
    category = db.Column(db.String(50), default='时间类计划')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    details = db.relationship('TimePlanDetail', backref='plan', lazy=True)  # 添加关联关系

class ProjectPlanDetail(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    plan_id = db.Column(db.Integer, db.ForeignKey('project_plan.id'), nullable=False)
    main_content = db.Column(db.String(200), nullable=False)
    key_node = db.Column(db.String(100))
    responsible_dept = db.Column(db.String(50))
    responsible_person = db.Column(db.String(20))
    cooperating_dept = db.Column(db.String(200))
    cooperating_persons = db.Column(db.String(200))
    responsible_leader = db.Column(db.String(20))
    deadline = db.Column(db.Date)
    completion_time = db.Column(db.Date)
    status = db.Column(db.String(10), default='进行中')
    last_status_update = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None))

class PlanProgressRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    detail_id = db.Column(db.Integer, db.ForeignKey('project_plan_detail.id'), nullable=False)
    submitter = db.Column(db.String(20), nullable=False)
    content = db.Column(db.Text, nullable=False)
    submit_time = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(10), default='待审核')
    reviewer = db.Column(db.String(20))

class OtherPlan(db.Model):
    __tablename__ = 'other_plan'
    id = db.Column(db.Integer, primary_key=True)
    plan_name = db.Column(db.String(100), nullable=False)
    category = db.Column(db.String(50), default='其他类计划')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    details = db.relationship('OtherPlanDetail', backref='plan', lazy='dynamic', cascade="all, delete-orphan")

    def to_dict(self):
        return {
            'id': self.id,
            'plan_name': self.plan_name,
            'category': self.category,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'details': [detail.to_dict() for detail in self.details.all()]  # 添加.all()确保加载关联数据
        }

class OtherPlanDetail(db.Model):
    __tablename__ = 'other_plan_detail'
    id = db.Column(db.Integer, primary_key=True)
    plan_id = db.Column(db.Integer, db.ForeignKey('other_plan.id'))
    content = db.Column(db.Text)
    status = db.Column(db.String(20))
    operators = db.Column(db.String(200))
    deadline = db.Column(db.Date)
    completion_time = db.Column(db.Date)
    
    def to_dict(self):
        try:
            content_data = json.loads(self.content) if self.content else {}
        except json.JSONDecodeError:
            content_data = {}
            
        return {
            'id': self.id,
            'content': content_data,
            'status': self.status,
            'operators': self.operators.split(',') if self.operators else [],
            'deadline': self.deadline.isoformat() if self.deadline else None,
            'completion_time': self.completion_time.isoformat() if self.completion_time else None,
            'plan_id': self.plan_id  # 添加关联ID字段
        }

class PersonalWeeklyReport(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    current_work = db.Column(db.Text)
    next_plan = db.Column(db.Text)
    submitted = db.Column(db.Boolean, default=False)
    department_submitted = db.Column(db.Boolean, default=False)
    submit_time = db.Column(db.DateTime)
    archived = db.Column(db.Boolean, default=False)
    report_date = db.Column(db.Date)
    user = db.relationship('User', backref='weekly_reports')

class PersonalWeeklyTemplate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    current_work = db.Column(db.Text)
    next_plan = db.Column(db.Text)

class HistoricalWeeklyReport(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    content = db.Column(db.Text)
    page = db.Column(db.Integer)

class CompanyWeeklyReport(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    department = db.Column(db.String(50), nullable=False)
    current_work = db.Column(db.Text)
    next_plan = db.Column(db.Text)
    submitted = db.Column(db.Boolean, default=False)
    submitter = db.Column(db.String(20))
    submit_time = db.Column(db.DateTime)
    archived = db.Column(db.Boolean, default=False)
    archive_time = db.Column(db.DateTime)

class DepartmentWeeklyReport(db.Model):
    __tablename__ = 'department_weekly_report'
    id = db.Column(db.Integer, primary_key=True)
    department = db.Column(db.String(50), nullable=False)
    report_content = db.Column(db.Text)
    submitter = db.Column(db.String(20))
    submit_date = db.Column(db.DateTime, default=datetime.utcnow)
    last_modified = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class TimePlanDetail(db.Model):
    __tablename__ = 'time_plan_detail'
    id = db.Column(db.Integer, primary_key=True)
    plan_id = db.Column(db.Integer, db.ForeignKey('time_plan.id'), nullable=False)
    plan_node = db.Column(db.String(200))
    responsible_leader = db.Column(db.String(50))
    responsible_dept = db.Column(db.String(50))
    cooperating_dept = db.Column(db.String(200))
    expected_deadline = db.Column(db.Date)
    actual_completion = db.Column(db.Date)
    progress = db.Column(db.Text)
    issues = db.Column(db.Text)
    status = db.Column(db.String(10), default='进行中')  # 新增状态字段
    operators = db.Column(db.String(200))
    bg_color = db.Column(db.String(20), default='#FFFFFF')

    def __repr__(self):
        return f'<TimePlanDetail {self.plan_node}>'

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls'}

def calculate_status(data):
    if not data['completion_time']:
        if data['deadline'] and data['deadline'] < datetime.now().date():
            return '逾期'
        return '进行中'
    else:
        if data['deadline'] and data['completion_time'] > data['deadline']:
            return '逾期完成'
        return '已完成'

scheduler = BackgroundScheduler(timezone="Asia/Shanghai")

def check_overdue_projects():
    with app.app_context():
        try:
            tz = ZoneInfo("Asia/Shanghai")
            now = datetime.now(tz).date()
            overdue_details = SupervisionDetail.query.filter(
                SupervisionDetail.status.in_(['进行中', '逾期']),
                SupervisionDetail.deadline < now
            ).all()
            
            for detail in overdue_details:
                if detail.status != '逾期':
                    detail.status = '逾期'
                    detail.last_status_update = datetime.now(tz)
            
            db.session.commit()
        except Exception as e:
            app.logger.error(f"定时任务执行错误: {str(e)}")

def transfer_weekly_reports():
    with app.app_context():
        try:
            tz = ZoneInfo("Asia/Shanghai")
            now = datetime.now(tz)
            
            personal_reports = WeeklyReport.query.filter_by(report_type='personal').all()
            for report in personal_reports:
                report.report_type = 'history_personal'
                report.report_date = now.date()
            
            department_reports = WeeklyReport.query.filter_by(report_type='department').all()
            for report in department_reports:
                new_history = WeeklyReport(
                    reporter=report.reporter,
                    department=report.department,
                    content=report.content,
                    report_type='history_department',
                    submit_time=now
                )
                db.session.add(new_history)
                db.session.delete(report)
            
            company_reports = CompanyWeeklyReport.query.all()
            for report in company_reports:
                new_history = WeeklyReport(
                    reporter='system',
                    department=report.department,
                    content=json.dumps({
                        'current_work': report.current_work,
                        'next_plan': report.next_plan
                    }),
                    report_type='history_company',
                    submit_time=now
                )
                db.session.add(new_history)
                db.session.delete(report)
            
            db.session.commit()
        except Exception as e:
            app.logger.error(f"周报转移错误: {str(e)}")

scheduler.add_job(check_overdue_projects, 'cron', hour=0)
scheduler.add_job(transfer_weekly_reports, 'cron', day_of_week='sun', hour=12)
scheduler.start()

# -------------------- 路由 --------------------
@app.route('/', methods=['GET', 'POST'])  
def login():
    if request.method == 'POST':
        user = User.query.filter_by(
            username=request.form['username'],
            password=request.form['password']
        ).first()
        if user:
            session['user'] = user.username
            session['is_admin'] = user.is_admin
            return redirect('/home')
    return render_template('login.html')

@app.route('/user_profile', methods=['GET', 'POST'])
def user_profile():
    if 'user' not in session:
        return redirect('/')
    
    user = User.query.filter_by(username=session['user']).first()
    
    if request.method == 'POST':
        user.phone = request.form['phone']
        user.password = request.form['password']
        db.session.commit()
        return redirect('/home')
    
    return render_template('user_profile.html',
                         username=user.username,
                         phone=user.phone,
                         password=user.password)

@app.route('/home')
def home():
    if 'user' not in session:
        return redirect('/')
    return render_template('index.html', 
                         is_admin=session.get('is_admin', False),
                         username=session['user'])

@app.route('/plan_management')
def plan_management():
    if 'user' not in session:
        return redirect('/')
    project_plans = ProjectPlan.query.all()
    time_plans = TimePlan.query.all()
    other_plans = OtherPlan.query.all()  
    return render_template('plan_management.html',
                         project_plans=project_plans,
                         time_plans=time_plans,
                         other_plans=other_plans)

@app.route('/add_project_plan', methods=['POST'])
def add_project_plan():
    if 'user' not in session or not session.get('is_admin'):
        return redirect('/')
    new_plan = ProjectPlan(plan_name=request.form['plan_name'])
    db.session.add(new_plan)
    db.session.commit()
    return redirect('/plan_management')

@app.route('/add_time_plan', methods=['POST'])
def add_time_plan():
    if 'user' not in session or not session.get('is_admin'):
        return redirect('/')
    new_plan = TimePlan(plan_name=request.form['plan_name'])
    db.session.add(new_plan)
    db.session.commit()
    return redirect('/plan_management')

@app.route('/delete_project_plan/<int:plan_id>')
def delete_project_plan(plan_id):
    if 'user' not in session or not session.get('is_admin'):
        return redirect('/')
    plan = ProjectPlan.query.get(plan_id)
    db.session.delete(plan)
    db.session.commit()
    return redirect('/plan_management')

@app.route('/delete_time_plan/<int:plan_id>')
def delete_time_plan(plan_id):
    if 'user' not in session or not session.get('is_admin'):
        return redirect('/')
    plan = TimePlan.query.get(plan_id)
    db.session.delete(plan)
    db.session.commit()
    return redirect('/plan_management')

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/')

@app.route('/user_management')
def user_management():
    if 'user' not in session or not session.get('is_admin'):
        return redirect('/')
    users = User.query.all()
    return render_template('user_management.html', users=users)

@app.route('/import_users', methods=['POST'])
def import_users():
    if 'user' not in session or not session.get('is_admin'):
        return redirect('/')
    
    if 'file' not in request.files:
        return "未选择文件", 400
    
    file = request.files['file']
    if file.filename == '':
        return "无效文件", 400
    
    try:
        df = pd.read_excel(file)
        for _, row in df.iterrows():
            new_user = User(
                username=row['用户名'],
                password=row['密码'],
                department=row.get('部门', ''),
                phone=row.get('电话', ''),
                is_admin=row.get('管理员', False),
                is_department_info=row.get('部门信息员', False),
                is_department_head=row.get('部门负责人', False),
                is_company_info=row.get('公司信息员', False),
                is_general_dept_head=row.get('综合部负责人', False),
                is_company_leader=row.get('公司领导', False)
            )
            db.session.add(new_user)
        db.session.commit()
        return redirect('/user_management')
    except Exception as e:
        db.session.rollback()
        return f"导入失败: {str(e)}", 400

@app.route('/export_users')
def export_users():
    if 'user' not in session or not session.get('is_admin'):
        return redirect('/')
    
    users = User.query.all()
    data = [{
        '用户名': u.username,
        '密码': u.password,
        '部门': u.department,
        '电话': u.phone,
        '管理员': u.is_admin,
        '部门信息员': u.is_department_info,
        '部门负责人': u.is_department_head,
        '公司信息员': u.is_company_info,
        '综合部负责人': u.is_general_dept_head,
        '公司领导': u.is_company_leader
    } for u in users]

    df = pd.DataFrame(data)
    output = BytesIO()
    df.to_excel(output, index=False, engine='openpyxl')
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name='用户列表.xlsx',
        as_attachment=True
    )

@app.route('/add_user', methods=['GET', 'POST'])
def add_user():
    if request.method == 'POST':
        new_user = User(
            username=request.form['username'],
            password=request.form['password'],
            department=request.form['department'],
            phone=request.form['phone'],
            is_admin='is_admin' in request.form,
            is_department_info='is_department_info' in request.form,
            is_department_head='is_department_head' in request.form,
            is_company_info='is_company_info' in request.form,
            is_general_dept_head='is_general_dept_head' in request.form,
            is_company_leader='is_company_leader' in request.form
        )
        db.session.add(new_user)
        db.session.commit()
        return redirect('/user_management')
    return render_template('add_user.html')

@app.route('/delete_user/<int:user_id>')
def delete_user(user_id):
    user = User.query.get(user_id)
    PersonalWeeklyReport.query.filter_by(user_id=user_id).delete()
    
    db.session.delete(user)
    db.session.commit()
    return redirect('/user_management')

@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
def edit_user(user_id):
    user = User.query.get(user_id)
    if request.method == 'POST':
        user.username = request.form['username']
        user.department = request.form.get('department', '')
        user.phone = request.form.get('phone', '')
        user.is_admin = 'is_admin' in request.form
        user.is_department_info = 'is_department_info' in request.form
        user.is_department_head = 'is_department_head' in request.form
        user.is_company_info = 'is_company_info' in request.form
        user.is_general_dept_head = 'is_general_dept_head' in request.form
        user.is_company_leader = 'is_company_leader' in request.form
        if request.form.get('password'):
            user.password = request.form['password']   
        try:
            db.session.commit()
            return redirect('/user_management')
        except Exception as e:
            db.session.rollback()
            return f"更新失败: {str(e)}", 500
    return render_template('edit_user.html', user=user)

@app.route('/supervision')
def supervision():
    if 'user' not in session:
        return redirect('/')
    projects = SupervisionProject.query.all()
    return render_template('supervision.html', projects=projects)

@app.route('/add_supervision', methods=['POST'])
def add_supervision():
    if 'user' not in session:
        return redirect('/')
    project_name = request.form['project_name']
    new_project = SupervisionProject(project_name=project_name)
    db.session.add(new_project)
    db.session.commit()
    return redirect('/supervision')

@app.route('/delete_supervision/<int:project_id>')
def delete_supervision(project_id):
    project = SupervisionProject.query.get(project_id)
    db.session.delete(project)
    db.session.commit()
    return redirect('/supervision')

@app.route('/supervision_detail/<int:project_id>')
def supervision_detail(project_id):
    if 'user' not in session:
        return redirect('/')
    
    project = db.session.get(SupervisionProject, project_id)  # 修改为新的session.get方法
    details = SupervisionDetail.query.filter_by(project_id=project_id).order_by(SupervisionDetail.main_content).all()
    
    merged_details = []
    prev_content = None
    for detail in details:
        if detail.main_content != prev_content:
            merged_details.append({
                'main_content': detail.main_content,
                'details': [detail],
                'rowspan': 1
            })
            prev_content = detail.main_content
        else:
            merged_details[-1]['details'].append(detail)
            merged_details[-1]['rowspan'] += 1
    
    users = User.query.all()
    leaders = User.query.filter_by(is_company_leader=True).all()
    
    return render_template('supervision_detail.html',
                         project=project,
                         merged_details=merged_details,
                         users=users,
                         leaders=leaders,
                         is_admin=session.get('is_admin', False),
                         source='supervision')

@app.route('/export_project_details/<int:project_id>')
def export_project_details(project_id):
    if 'user' not in session:
        return redirect('/')
    
    details = SupervisionDetail.query.filter_by(project_id=project_id).all()
    
    data = [{
        '主要内容': d.main_content,
        '关键节点': d.key_node or '',
        '责任部门': d.responsible_dept or '',
        '责任人': d.responsible_person or '',
        '配合部门': d.cooperating_dept or '',
        '配合人': d.cooperating_persons or '',
        '责任领导': d.responsible_leader or '',
        '完成时限': d.deadline.strftime('%Y-%m-%d') if d.deadline else '',
        '完成时间': d.completion_time.strftime('%Y-%m-%d') if d.completion_time else '',
        '状态': d.status or '进行中'
    } for d in details]

    df = pd.DataFrame(data)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='项目详情')
        worksheet = writer.sheets['项目详情']
        worksheet.column_dimensions['A'].width = 30  # 主要内容
        worksheet.column_dimensions['B'].width = 15  # 关键节点
        worksheet.column_dimensions['C'].width = 15  # 责任部门
        worksheet.column_dimensions['D'].width = 10  # 责任人
        worksheet.column_dimensions['E'].width = 15  # 配合部门
        worksheet.column_dimensions['F'].width = 15  # 配合人
        worksheet.column_dimensions['G'].width = 10  # 责任领导
        worksheet.column_dimensions['H'].width = 12  # 完成时限
        worksheet.column_dimensions['I'].width = 12  # 完成时间
        worksheet.column_dimensions['J'].width = 10  # 状态
    
    output.seek(0)
    filename = f"project_{project_id}_details.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name=filename,
        as_attachment=True
    )

@app.route('/add_project_detail/<int:project_id>', methods=['POST'])
def add_project_detail(project_id):
    if 'user' not in session or not session.get('is_admin'):
        return redirect('/')
    
    try:
        deadline_str = request.form.get('deadline')
        deadline = None
        if deadline_str:
            deadline = datetime.strptime(deadline_str, '%Y-%m-%d').date()
        
        new_detail = SupervisionDetail(
            project_id=project_id,
            main_content=request.form['main_content'],
            key_node=request.form['key_node'],
            responsible_dept=request.form['responsible_dept'],
            responsible_person=request.form['responsible_person'],
            cooperating_dept=request.form.get('cooperating_dept', ''),
            cooperating_persons=','.join(request.form.getlist('cooperating_persons')),
            responsible_leader=request.form['responsible_leader'],
            deadline=deadline,  # 使用处理后的日期对象
            status='进行中'
        )
        
        if new_detail.deadline and new_detail.deadline < datetime.now().date():
            new_detail.status = '逾期'
        
        db.session.add(new_detail)
        db.session.commit()
        return redirect(f'/supervision_detail/{project_id}')
    except ValueError as e:
        db.session.rollback()
        print(f"日期格式错误: {str(e)}")
        return "日期格式无效，请使用YYYY-MM-DD格式", 400
    except Exception as e:
        db.session.rollback()
        print(f"添加失败: {str(e)}")
        return f"提交失败: {str(e)}", 500

@app.route('/complete_project/<int:detail_id>', methods=['POST'])
def complete_project(detail_id):
    if 'user' not in session or not session.get('is_admin'):
        return redirect('/')
    
    try:
        completion_time_str = request.form.get('completion_time', '').strip()
        if not completion_time_str:
            raise ValueError("缺失完成时间参数")
        try:
            completion_time = datetime.strptime(completion_time_str, '%Y-%m-%d').date()
        except ValueError:
            raise ValueError("无效的日期格式，必须为YYYY-MM-DD")
        
        detail = SupervisionDetail.query.with_for_update().get_or_404(detail_id)
        detail.completion_time = completion_time
        if detail.deadline:
            if completion_time > detail.deadline:
                detail.status = '逾期完成'
            else:
                detail.status = '已完成'
        else:
            detail.status = '已完成'
            
        detail.last_status_update = datetime.utcnow()
        db.session.commit()
        
        return redirect(url_for('supervision_detail', project_id=detail.project_id))
        
    except ValueError as e:
        db.session.rollback()
        flash(f'操作失败: {str(e)}', 'error')
        return redirect(url_for('supervision_detail', project_id=detail.project_id))
    
    except SQLAlchemyError as e:  
        db.session.rollback()
        app.logger.error(f"数据库操作异常: {str(e)}")
        flash('操作失败：数据库错误', 'error')
        return redirect(url_for('supervision_detail', project_id=detail.project_id))
    
    except Exception as e:
        db.session.rollback()
        app.logger.critical(f"系统异常: {str(e)}")
        flash('操作失败：系统错误', 'error')
        return redirect(url_for('supervision_detail', project_id=detail.project_id))

@app.route('/import_project_details/<int:project_id>', methods=['POST'])
def import_project_details(project_id):
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({"error": "Unauthorized operation"}), 401

    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        merged_values = {}
        for merged_range in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            top_left_value = ws.cell(row=min_row, column=min_col).value
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged_values[(row, col)] = top_left_value

        headers = []
        for cell in ws[1]:
            header = str(cell.value).strip() if cell.value else f"列{cell.column_letter}"
            headers.append(header)

        imported_data = []
        for row in ws.iter_rows(min_row=2):
            if all(cell.value is None for cell in row):
                continue

            item = {}
            for col_idx, cell in enumerate(row):
                if col_idx >= len(headers):
                    continue
                    
                header = headers[col_idx]
                cell_value = merged_values.get((cell.row, cell.column), cell.value)
                
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.date()
                elif isinstance(cell_value, (int, float)) and not isinstance(cell_value, bool):
                    cell_value = cell_value
                else:
                    cell_value = str(cell_value).strip() if cell_value not in (None, "") else ""

                item[header] = cell_value

            # 关键节点日期提取（保留原始内容）
            key_node = item.get('关键节点', '')
            deadline = None
            if key_node:
                date_match = re.search(
                    r'(\d{4})[年/-]?(\d{1,2})?[月/-]?(\d{1,2})?',
                    key_node
                )
                if date_match:
                    try:
                        year = int(date_match.group(1))
                        month = int(date_match.group(2)) if date_match.group(2) else 1
                        day = int(date_match.group(3)) if date_match.group(3) else 1
                        deadline = date(year, month, day)
                    except ValueError as e:
                        app.logger.warning(f"日期解析失败: {key_node}，错误: {str(e)}")
                        deadline = None

            # 状态计算
            completion = None
            if item.get('完成时间'):
                try:
                    completion = datetime.strptime(item['完成时间'], '%Y-%m-%d').date()
                except ValueError:
                    pass

            current_date = datetime.now().date()
            status = '进行中'
            if completion:
                status = '逾期完成' if (deadline and completion > deadline) else '已完成'
            elif deadline:
                status = '逾期' if deadline < current_date else '进行中'

            imported_data.append({
                'main_content': item.get('主要内容', '未命名任务'),
                'key_node': key_node,  # 保留原始关键节点内容
                'responsible_dept': item.get('责任部门', ''),
                'responsible_person': item.get('责任人', ''),
                'cooperating_dept': item.get('配合部门', ''),
                'cooperating_persons': ','.join(item.get('配合人', '').split(',')),
                'responsible_leader': item.get('责任领导', ''),
                'deadline': deadline,
                'completion_time': completion,
                'status': status
            })

        try:
            db.session.begin()
            SupervisionDetail.query.filter_by(project_id=project_id).delete()
            
            for item in imported_data:
                new_detail = SupervisionDetail(
                    project_id=project_id,
                    main_content=item['main_content'],
                    key_node=item['key_node'],
                    responsible_dept=item['responsible_dept'],
                    responsible_person=item['responsible_person'],
                    cooperating_dept=item['cooperating_dept'],
                    cooperating_persons=item['cooperating_persons'],
                    responsible_leader=item['responsible_leader'],
                    deadline=item['deadline'],
                    completion_time=item['completion_time'],
                    status=item['status']
                )
                db.session.add(new_detail)
            
            db.session.commit()
            return jsonify({
                "message": f"成功导入{len(imported_data)}条数据",
                "redirect": f"/supervision_detail/{project_id}"
            }), 200

        except SQLAlchemyError as e:
            db.session.rollback()
            app.logger.error(f"数据库错误: {str(e)}")
            return jsonify({"error": f"数据库操作失败: {str(e)}"}), 500
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"数据处理错误: {str(e)}")
            return jsonify({"error": f"数据处理失败: {str(e)}"}), 500

    except Exception as e:
        app.logger.error(f"文件处理失败: {str(e)}", exc_info=True)
        return jsonify({"error": f"文件处理失败: {str(e)}"}), 500

@app.route('/update_project_deadline/<int:detail_id>', methods=['POST'])
def update_project_deadline(detail_id):
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({'status': 'error', 'message': '未授权操作'}), 401
    
    try:
        detail = SupervisionDetail.query.get_or_404(detail_id)
        new_deadline = request.json.get('deadline')
        
        if new_deadline:
            detail.deadline = datetime.strptime(new_deadline, '%Y-%m-%d').date()
            # 自动更新状态
            if detail.completion_time:
                detail.status = '逾期完成' if detail.completion_time > detail.deadline else '已完成'
            else:
                detail.status = '逾期' if detail.deadline < datetime.now().date() else '进行中'
        
        db.session.commit()
        return jsonify({'status': 'success'})
    
    except ValueError:
        return jsonify({'status': 'error', 'message': '无效日期格式'}), 400
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"更新时限失败: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/delete_project_detail/<int:detail_id>', methods=['DELETE'])
def delete_project_detail(detail_id):
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({'status': 'error', 'message': '未授权操作'}), 401
    
    try:
        detail = SupervisionDetail.query.get_or_404(detail_id)
        project_id = detail.project_id
        db.session.delete(detail)
        db.session.commit()
        return jsonify({
            'status': 'success',
            'redirect': f'/supervision_detail/{project_id}'
        })
    except SQLAlchemyError as e:
        db.session.rollback()
        app.logger.error(f"删除督办条目失败: {str(e)}")
        return jsonify({'status': 'error', 'message': '数据库操作失败'}), 500
    except Exception as e:
        app.logger.error(f"系统异常: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/submit_progress/<int:detail_id>', methods=['POST'])
def submit_progress(detail_id):
    if 'user' not in session:
        return redirect('/')
    
    detail = SupervisionDetail.query.get(detail_id)
    current_user = session['user']
    
    if current_user not in [detail.responsible_person] + detail.cooperating_persons.split(','):
        return "无操作权限", 403
    
    new_progress = ProgressRecord(
        detail_id=detail_id,
        submitter=current_user,
        content=request.form['content']
    )
    db.session.add(new_progress)
    db.session.commit()
    return redirect(f'/project_progress/{detail_id}')

@app.route('/review_progress/<int:progress_id>/<action>')
def review_progress(progress_id, action):
    if 'user' not in session or not session.get('is_admin'):
        return redirect('/')
    
    progress = ProgressRecord.query.get(progress_id)
    progress.status = '已通过' if action == 'approve' else '已驳回'
    progress.reviewer = session['user']
    db.session.commit()
    return redirect(f'/project_progress/{progress.detail_id}')

@app.route('/project_progress/<int:detail_id>')
def project_progress(detail_id):
    if 'user' not in session:
        return redirect('/')
    
    # 获取督办详情
    detail = db.session.get(SupervisionDetail, detail_id)
    if not detail:
        abort(404)
    
    # 获取当前用户（根据用户名查询）
    current_user = User.query.filter_by(username=session['user']).first()
    if not current_user:
        abort(401, description="用户不存在")
    
    # 处理协作者信息
    cooperating_persons = []
    if detail.cooperating_persons:
        cooperating_persons = [cp.strip() for cp in detail.cooperating_persons.split(',')]
    
    # 权限验证
    has_access = (
        current_user.username == detail.responsible_person or
        current_user.username in cooperating_persons
    )
    
    # 获取进度记录
    progress_records = ProgressRecord.query.filter_by(detail_id=detail_id).order_by(ProgressRecord.submit_time.desc()).all()
    
    return render_template('project_progress.html',
                         detail=detail,
                         progress_records=progress_records,
                         is_admin=current_user.is_admin,
                         is_company_leader=current_user.is_company_leader,
                         has_access=has_access,
                         current_user=current_user,
                         source='supervision')  # 添加source参数区分督办类

@app.route('/plan_progress/<int:detail_id>')
def plan_progress(detail_id):
    if 'user' not in session:
        return redirect('/')
    
    # 获取计划详情
    detail = db.session.get(ProjectPlanDetail, detail_id)
    if not detail:
        abort(404)
    
    # 获取当前用户（根据用户名查询）
    current_user = User.query.filter_by(username=session['user']).first()
    if not current_user:
        abort(401, description="用户不存在")
    
    # 处理协作者信息
    cooperating_persons = []
    if detail.cooperating_persons:
        cooperating_persons = [cp.strip() for cp in detail.cooperating_persons.split(',')]
    
    # 权限验证
    has_access = (
        current_user.username == detail.responsible_person or
        current_user.username in cooperating_persons
    )
    
    # 获取进度记录
    progress_records = ProgressRecord.query.filter_by(detail_id=detail_id).order_by(ProgressRecord.submit_time.desc()).all()
    
    return render_template('project_progress.html',
                         detail=detail,
                         progress_records=progress_records,
                         is_admin=current_user.is_admin,
                         is_company_leader=current_user.is_company_leader,
                         has_access=has_access,
                         current_user=current_user,
                         source='plan')

@app.route('/project_plan_detail/<int:plan_id>')
def project_plan_detail(plan_id):
    if 'user' not in session:
        return redirect('/')
    
    plan = ProjectPlan.query.get_or_404(plan_id)
    details = ProjectPlanDetail.query.filter_by(plan_id=plan_id).order_by(ProjectPlanDetail.main_content).all()
    merged_details = []
    prev_content = None
    for detail in details:
        if detail.main_content != prev_content:
            merged_details.append({
                'main_content': detail.main_content,
                'details': [detail],
                'rowspan': 1
            })
            prev_content = detail.main_content
        else:
            merged_details[-1]['details'].append(detail)
            merged_details[-1]['rowspan'] += 1
    
    users = User.query.all()
    leaders = User.query.filter_by(is_company_leader=True).all()
    
    return render_template('supervision_detail.html',
                         project=plan,
                         merged_details=merged_details,
                         users=users,
                         leaders=leaders,
                         is_admin=session.get('is_admin', False),
                         source='plan')

@app.route('/delete_plan_detail/<int:detail_id>', methods=['DELETE'])
def delete_plan_detail(detail_id):
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({'status': 'error', 'message': '未授权操作'}), 401
    
    try:
        detail = ProjectPlanDetail.query.get_or_404(detail_id)
        plan_id = detail.plan_id
        db.session.delete(detail)
        db.session.commit()
        return jsonify({
            'status': 'success',
            'redirect': f'/project_plan_detail/{plan_id}'
        })
    except SQLAlchemyError as e:
        db.session.rollback()
        app.logger.error(f"删除计划条目失败: {str(e)}")
        return jsonify({'status': 'error', 'message': '数据库操作失败'}), 500
    except Exception as e:
        app.logger.error(f"系统异常: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/add_plan_detail/<int:plan_id>', methods=['POST'])
def add_plan_detail(plan_id):
    if 'user' not in session or not session.get('is_admin'):
        return redirect('/')
    
    try:
        deadline_str = request.form.get('deadline')
        deadline = None
        if deadline_str:
            deadline = datetime.strptime(deadline_str, '%Y-%m-%d').date()
        
        new_detail = ProjectPlanDetail(
            plan_id=plan_id,
            main_content=request.form['main_content'],
            key_node=request.form['key_node'],
            responsible_dept=request.form['responsible_dept'],
            responsible_person=request.form['responsible_person'],
            cooperating_dept=request.form.get('cooperating_dept', ''),
            cooperating_persons=','.join(request.form.getlist('cooperating_persons')),
            responsible_leader=request.form['responsible_leader'],
            deadline=deadline,
            status='进行中'
        )
        
        if new_detail.deadline and new_detail.deadline < datetime.now().date():
            new_detail.status = '逾期'
        
        db.session.add(new_detail)
        db.session.commit()
        return redirect(f'/project_plan_detail/{plan_id}')
    except ValueError as e:
        db.session.rollback()
        print(f"日期格式错误: {str(e)}")
        return "日期格式无效，请使用YYYY-MM-DD格式", 400
    except Exception as e:
        db.session.rollback()
        print(f"添加失败: {str(e)}")
        return f"提交失败: {str(e)}", 500

@app.route('/complete_plan_detail/<int:detail_id>', methods=['POST'])
def complete_plan_detail(detail_id):
    if 'user' not in session or not session.get('is_admin'):
        return redirect('/')
    
    try:
        completion_time_str = request.form.get('completion_time', '').strip()
        if not completion_time_str:
            raise ValueError("缺失完成时间参数")
        try:
            completion_time = datetime.strptime(completion_time_str, '%Y-%m-%d').date()
        except ValueError:
            raise ValueError("无效的日期格式，必须为YYYY-MM-DD")
        
        detail = ProjectPlanDetail.query.with_for_update().get_or_404(detail_id)
        detail.completion_time = completion_time
        if detail.deadline:
            if completion_time > detail.deadline:
                detail.status = '逾期完成'
            else:
                detail.status = '已完成'
        else:
            detail.status = '已完成'
            
        detail.last_status_update = datetime.utcnow()
        db.session.commit()
        
        return redirect(url_for('project_plan_detail', plan_id=detail.plan_id))
        
    except ValueError as e:
        db.session.rollback()
        flash(f'操作失败: {str(e)}', 'error')
        return redirect(url_for('project_plan_detail', plan_id=detail.plan_id))
    
    except SQLAlchemyError as e:  
        db.session.rollback()
        app.logger.error(f"数据库操作异常: {str(e)}")
        flash('操作失败：数据库错误', 'error')
        return redirect(url_for('project_plan_detail', plan_id=detail.plan_id))
    
    except Exception as e:
        db.session.rollback()
        app.logger.critical(f"系统异常: {str(e)}")
        flash('操作失败：系统错误', 'error')
        return redirect(url_for('project_plan_detail', plan_id=detail.plan_id))

@app.route('/update_plan_deadline/<int:detail_id>', methods=['POST'])
def update_plan_deadline(detail_id):
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({'status': 'error', 'message': '未授权操作'}), 401
    
    try:
        detail = ProjectPlanDetail.query.get_or_404(detail_id)
        new_deadline = request.json.get('deadline')
        
        if new_deadline:
            detail.deadline = datetime.strptime(new_deadline, '%Y-%m-%d').date()
            if detail.completion_time:
                detail.status = '逾期完成' if detail.completion_time > detail.deadline else '已完成'
            else:
                detail.status = '逾期' if detail.deadline < datetime.now().date() else '进行中'
        
        db.session.commit()
        return jsonify({'status': 'success'})
    
    except ValueError:
        return jsonify({'status': 'error', 'message': '无效日期格式'}), 400
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"更新计划时限失败: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/time_plan/<int:plan_id>')
def time_plan(plan_id):
    if 'user' not in session:
        return redirect('/')
    
    current_user = User.query.filter_by(username=session['user']).first()
    details = TimePlanDetail.query.filter_by(plan_id=plan_id).order_by(TimePlanDetail.id).all()
    
    departments = [u.department for u in User.query.group_by(User.department).all()]
    
    is_responsible = any(d.responsible_dept == current_user.department for d in details)
    can_edit = current_user.is_admin or is_responsible
    
    return render_template('time_plan.html',
                         details=details,
                         can_edit=can_edit,
                         is_admin=current_user.is_admin,
                         departments=departments,
                         current_user=current_user)

@app.route('/update_time_plan/<int:detail_id>', methods=['POST'])
def update_time_plan(detail_id):
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({'status': 'error', 'message': '未授权操作'}), 401
    try:
        detail = TimePlanDetail.query.get_or_404(detail_id)
        if 'expected_deadline' in request.form:
            new_date = request.form['expected_deadline']
            detail.expected_deadline = datetime.strptime(new_date, '%Y-%m-%d').date() if new_date else None
        if 'actual_completion' in request.form:
            completion_date = request.form['actual_completion']
            detail.actual_completion = datetime.strptime(completion_date, '%Y-%m-%d').date() if completion_date else None
        if 'progress' in request.form:
            detail.progress = request.form['progress']
        if 'issues' in request.form:
            detail.issues = request.form['issues']
        current_date = datetime.now().date()
        if detail.actual_completion:
            if detail.expected_deadline:
                detail.status = '逾期完成' if detail.actual_completion > detail.expected_deadline else '已完成'
            else:
                detail.status = '已完成'
        elif detail.expected_deadline:
            detail.status = '逾期' if detail.expected_deadline < current_date else '进行中'
        else:
            detail.status = '进行中'
        
        db.session.commit()
        return jsonify({'status': 'success'})
    
    except ValueError as e:
        db.session.rollback()
        app.logger.error(f"日期格式错误: {str(e)}")
        return jsonify({'status': 'error', 'message': '无效的日期格式，请使用YYYY-MM-DD'}), 400
    except SQLAlchemyError as e:
        db.session.rollback()
        app.logger.error(f"数据库错误: {str(e)}")
        return jsonify({'status': 'error', 'message': '数据库操作失败'}), 500
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"系统异常: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/import_time_plan', methods=['POST'])
def import_time_plan():
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({'status': 'error', 'message': '未授权操作'}), 401

    if 'file' not in request.files:
        return jsonify({'status': 'error', 'message': '未上传文件'}), 400

    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'status': 'error', 'message': '未选择文件'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        # 处理合并单元格
        merged_values = {}
        for merged_range in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            top_left_value = ws.cell(row=min_row, column=min_col).value
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged_values[(row, col)] = top_left_value

        headers = []
        for cell in ws[1]:
            column_letter = get_column_letter(cell.column)
            header = str(cell.value).strip() if cell.value else f"列{column_letter}"
            headers.append(header)

        imported_data = []
        for row in ws.iter_rows(min_row=2):
            if all(cell.value is None for cell in row):
                continue

            item = {}
            for col_idx, cell in enumerate(row):
                if col_idx >= len(headers):
                    continue
                    
                header = headers[col_idx]
                cell_value = merged_values.get((cell.row, cell.column), cell.value)
                
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.date()
                elif isinstance(cell_value, (int, float)) and not isinstance(cell_value, bool):
                    cell_value = cell_value
                else:
                    cell_value = str(cell_value).strip() if cell_value not in (None, "") else ""

                item[header] = cell_value

            # 日期解析逻辑
            def parse_date(value):
                if isinstance(value, date):
                    return value
                try:
                    return datetime.strptime(value, '%Y-%m-%d').date()
                except (TypeError, ValueError):
                    date_match = re.search(r'(\d{4})[年/-]?(\d{1,2})[月/-]?(\d{1,2})日?', str(value))
                    if date_match:
                        year, month, day = map(int, date_match.groups())
                        return date(year, month, day)
                    return None

            expected_deadline = parse_date(item.get('预计完成时间'))
            actual_completion = parse_date(item.get('实际完成时间'))

            imported_data.append({
                'plan_node': item.get('集团下发月度计划节点', '未命名节点'),
                'responsible_leader': item.get('责任领导', ''),
                'responsible_dept': item.get('责任部门', ''),
                'cooperating_dept': item.get('配合部门', ''),
                'expected_deadline': expected_deadline,
                'actual_completion': actual_completion,
                'progress': item.get('进展及存在问题', ''),
                'issues': item.get('需协调解决事项', ''),
                'operators': ','.join(set(filter(None, [
                    *str(item.get('责任人', '')).split(','),
                    *str(item.get('责任领导', '')).split(','),
                    *str(item.get('配合人', '')).split(',')
                ]))),
                'bg_color': item.get('背景色', '#FFFFFF')
            })

        try:
            db.session.begin()
            
            # 处理计划关联
            current_plan = db.session.query(TimePlan).first()
            if not current_plan:
                current_plan = TimePlan(plan_name="默认时间计划")
                db.session.add(current_plan)
                db.session.flush()

            TimePlanDetail.query.filter_by(plan_id=current_plan.id).delete()
            
            for item in imported_data:
                new_detail = TimePlanDetail(
                    plan_id=current_plan.id,
                    plan_node=item['plan_node'],
                    responsible_leader=item['responsible_leader'],
                    responsible_dept=item['responsible_dept'],
                    cooperating_dept=item['cooperating_dept'],
                    expected_deadline=item['expected_deadline'],
                    actual_completion=item['actual_completion'],
                    progress=item['progress'],
                    issues=item['issues'],
                    operators=item['operators'],
                    bg_color=item['bg_color']
                )
                db.session.add(new_detail)
            
            db.session.commit()
            return jsonify({
                'status': 'success',
                'message': f'成功导入{len(imported_data)}条数据',
                'redirect': f'/time_plan/{current_plan.id}'
            }), 200

        except SQLAlchemyError as e:
            db.session.rollback()
            app.logger.error(f"数据库错误: {str(e)}", exc_info=True)
            return jsonify({'status': 'error', 'message': f'数据库操作失败: {str(e)}'}), 500
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"数据处理错误: {str(e)}", exc_info=True)
            return jsonify({'status': 'error', 'message': f'数据处理失败: {str(e)}'}), 500

    except Exception as e:
        app.logger.error(f"文件处理失败: {str(e)}", exc_info=True)
        return jsonify({'status': 'error', 'message': f'文件处理失败: {str(e)}'}), 500

@app.route('/export_time_plan')
def export_time_plan():
    if 'user' not in session:
        return redirect('/')
    
    try:
        # 获取所有时间计划数据
        details = TimePlanDetail.query.order_by(TimePlanDetail.id).all()
        
        # 构建导出数据
        data = [{
            '序号': idx + 1,
            '集团下发月度计划节点': d.plan_node,
            '责任领导': d.responsible_leader,
            '责任部门': d.responsible_dept,
            '配合部门': d.cooperating_dept,
            '预计完成时间': d.expected_deadline.strftime('%Y-%m-%d') if d.expected_deadline else '',
            '实际完成时间': d.actual_completion.strftime('%Y-%m-%d') if d.actual_completion else '',
            '进展情况': d.progress,
            '存在问题': d.issues,
            '当月是否完成': '是' if d.is_completed else '否'
        } for idx, d in enumerate(details)]

        df = pd.DataFrame(data)
        
        # 创建内存文件对象
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='时间类计划')
            
            # 获取工作表对象进行样式调整
            worksheet = writer.sheets['时间类计划']
            
            # 设置列宽（单位：字符）
            column_widths = {
                'A': 8,   # 序号
                'B': 35,  # 计划节点
                'C': 12,  # 责任领导
                'D': 15,  # 责任部门
                'E': 15,  # 配合部门
                'F': 15,  # 预计完成时间
                'G': 15,  # 实际完成时间
                'H': 40,  # 进展情况
                'I': 40,  # 存在问题
                'J': 12   # 是否完成
            }
            
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            # 设置标题行样式
            header_row = worksheet[1]
            for cell in header_row:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center')
                cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        output.seek(0)
        filename = f"time_plan_export_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            download_name=filename,
            as_attachment=True
        )
    
    except SQLAlchemyError as e:
        app.logger.error(f"数据库查询失败: {str(e)}")
        return "数据导出失败，请稍后重试", 500
    except Exception as e:
        app.logger.error(f"文件生成失败: {str(e)}")
        return "文件生成错误，请联系管理员", 500

@app.route('/delete_time_plan_row/<int:row_id>', methods=['DELETE'])
def delete_time_plan_row(row_id):
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({'status': 'error', 'message': '未授权操作'}), 401
    
    try:
        # 使用新的session.get方法替代旧的query.get
        detail = db.session.get(TimePlanDetail, row_id)
        if not detail:
            return jsonify({'status': 'error', 'message': '条目不存在'}), 404
            
        plan_id = detail.plan_id
        db.session.delete(detail)
        db.session.commit()
        return jsonify({
            'status': 'success',
            'message': '删除成功',
            'redirect': f'/time_plan/{plan_id}'
        })
    except SQLAlchemyError as e:
        db.session.rollback()
        app.logger.error(f"删除失败-数据库错误: {str(e)}")
        return jsonify({'status': 'error', 'message': f'数据库操作失败: {str(e)}'}), 500
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"删除失败-系统异常: {str(e)}", exc_info=True)
        return jsonify({'status': 'error', 'message': f'系统异常: {str(e)}'}), 500

@app.route('/other_plan_detail/<int:plan_id>')
def other_plan_detail(plan_id):
    if 'user' not in session:
        return redirect('/')
    
    plan = OtherPlan.query.get_or_404(plan_id)
    details = OtherPlanDetail.query.filter_by(plan_id=plan_id).order_by(OtherPlanDetail.id).all()
    users = [{'username': u.username} for u in User.query.all()]
    
    merged_data = []
    current_main_content = None
    current_group = None
    
    for detail in details:
        try:
            content = json.loads(detail.content)
        except json.JSONDecodeError:
            content = {}
        
        main_content = content.get('任务名称') or content.get('主要内容') or '未命名任务'
        
        if main_content != current_main_content:
            current_main_content = main_content
            current_group = {
                'id': detail.id,
                'main_content': main_content,
                'rows': [],
                'rowspan': 0
            }
            merged_data.append(current_group)
        
        row_data = {
            'id': detail.id,
            **content,
            'status': detail.status,
            'operators': detail.operators.split(',') if detail.operators else [],
            'deadline': detail.deadline.strftime('%Y-%m-%d') if detail.deadline else '',
            'completion_time': detail.completion_time.strftime('%Y-%m-%d') if detail.completion_time else ''
        }
        
        if not row_data['status']:
            deadline = detail.deadline.date() if detail.deadline else None
            completion = detail.completion_time.date() if detail.completion_time else None
            
            if completion:
                if deadline and completion > deadline:
                    row_data['status'] = '逾期完成'
                else:
                    row_data['status'] = '已完成'
            else:
                if deadline and deadline < datetime.now().date():
                    row_data['status'] = '逾期'
                else:
                    row_data['status'] = '进行中'
        
        current_group['rows'].append(row_data)
        current_group['rowspan'] += 1
    
    columns = []
    if merged_data:
        sample_row = merged_data[0]['rows'][0] if merged_data[0]['rows'] else {}
        columns = [key for key in sample_row.keys() 
                  if key not in ['id', 'status', 'operators', 'deadline', 'completion_time', '任务名称']]
    
    return render_template('other_plan_detail.html',
                         plan=plan,
                         table_data=merged_data,
                         columns=columns,
                         is_admin=session.get('is_admin', False),
                         current_user=session.get('user'),
                         users=users)

@app.route('/import_other_plan/<int:plan_id>', methods=['POST'])
def import_other_plan(plan_id):
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({"error": "Unauthorized operation"}), 401

    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        # 处理合并单元格
        merged_values = {}
        for merged_range in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            top_left_value = ws.cell(row=min_row, column=min_col).value
            
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged_values[(row, col)] = top_left_value

        # 动态处理表头
        headers = []
        for cell in ws[1]:
            header = str(cell.value).strip() if cell.value else f"列{cell.column_letter}"
            headers.append(header)

        imported_data = []
        
        # 增强的日期解析函数
        def parse_date(value):
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, str):
                try:
                    return datetime.strptime(value, '%Y-%m-%d').date()
                except ValueError:
                    pass
                match = re.search(
                    r'(\d{4})[年/-]?(\d{1,2})?[月/-]?(\d{1,2})?',
                    value
                )
                if match:
                    year = int(match.group(1)) if match.group(1) else datetime.now().year
                    month = int(match.group(2)) if match.group(2) else 1
                    day = int(match.group(3)) if match.group(3) else 1
                    try:
                        return date(year, month, day)
                    except ValueError:
                        app.logger.warning(f"无效日期格式: {value}")
            return None

        # 处理数据行
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if all(cell.value is None for cell in row):
                continue

            item = {}
            for col_idx, cell in enumerate(row):
                if col_idx >= len(headers):
                    continue
                    
                header = headers[col_idx]
                cell_value = merged_values.get((cell.row, cell.column), cell.value)
                
                # 处理数据类型
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.date()
                elif isinstance(cell_value, (int, float)) and not isinstance(cell_value, bool):
                    cell_value = cell_value  # 保留原始数值类型
                else:
                    cell_value = str(cell_value).strip() if cell_value not in (None, "") else ""

                # 处理特殊字段
                if header in ['完成时限', '完成时间']:
                    item[header] = parse_date(cell_value)
                else:
                    item[header] = cell_value

            # 动态生成主要内容
            main_content = item.get('任务名称') or item.get('主要内容') or '未命名任务'
            item['任务名称'] = main_content

            # 自动状态计算
            deadline = item.get('完成时限')
            completion = item.get('完成时间')
            current_date = datetime.now().date()
            
            status = item.get('状态', '')
            if not status:
                if completion:
                    status = '逾期完成' if (deadline and completion > deadline) else '已完成'
                else:
                    status = '逾期' if (deadline and deadline < current_date) else '进行中'
            
            # 操作人员处理
            operators = []
            if '操作权限' in item and item['操作权限']:
                operators = [op.strip() for op in str(item['操作权限']).split(',') if op.strip()]

            # 构建动态内容
            dynamic_content = {k: v for k, v in item.items() if k not in ['status', 'operators']}
            
            imported_data.append({
                'content': dynamic_content,
                'status': status,
                'operators': operators,
                'deadline': item.get('完成时限'),
                'completion_time': item.get('完成时间')
            })

        try:
            db.session.begin()
            
            OtherPlanDetail.query.filter_by(plan_id=plan_id).delete()
            
            for data in imported_data:
                new_detail = OtherPlanDetail(
                    plan_id=plan_id,
                    content=json.dumps(data['content'], ensure_ascii=False),
                    status=data['status'],
                    operators=','.join(data['operators']),
                    deadline=data['deadline'],
                    completion_time=data['completion_time']
                )
                db.session.add(new_detail)
            
            plan = OtherPlan.query.get(plan_id)
            if plan:
                plan.created_at = datetime.utcnow()
            
            db.session.commit()
            
            # 构建返回数据
            new_details = OtherPlanDetail.query.filter_by(plan_id=plan_id).all()
            result = [{
                'id': d.id,
                'content': json.loads(d.content),
                'status': d.status,
                'operators': d.operators.split(',') if d.operators else [],
                'deadline': d.deadline.isoformat() if d.deadline else None,
                'completion_time': d.completion_time.isoformat() if d.completion_time else None
            } for d in new_details]

            return jsonify({
                "status": "success",
                "message": f"成功导入{len(imported_data)}条数据",
                "data": result,
                "columns": list(imported_data[0]['content'].keys()) if imported_data else []
            }), 200

        except SQLAlchemyError as e:
            db.session.rollback()
            app.logger.error(f"数据库错误: {str(e)}")
            return jsonify({"status": "error", "message": f"数据库操作失败: {str(e)}"}), 500
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"数据处理错误: {str(e)}")
            return jsonify({"status": "error", "message": f"数据处理失败: {str(e)}"}), 500

    except openpyxl.utils.exceptions.InvalidFileException:
        return jsonify({"status": "error", "message": "无效的Excel文件格式"}), 400
    except Exception as e:
        app.logger.error(f"文件处理失败: {str(e)}", exc_info=True)
        return jsonify({"status": "error", "message": f"文件处理失败: {str(e)}"}), 500

@app.route('/add_other_plan', methods=['POST'])
def add_other_plan():
    if 'user' not in session or not session.get('is_admin'):
        return redirect('/')
    
    plan_name = request.form['plan_name']
    new_plan = OtherPlan(
        plan_name=plan_name,
        category='其他类计划',
        created_at=datetime.utcnow()
    )
    
    try:
        db.session.add(new_plan)
        db.session.commit()
        return redirect('/plan_management')
    except IntegrityError:
        db.session.rollback()
        flash('计划名称已存在', 'error')
        return redirect('/plan_management')
    except Exception as e:
        db.session.rollback()
        flash(f'创建失败: {str(e)}', 'error')
        return redirect('/plan_management')

@app.route('/export_other_plan/<int:plan_id>')
def export_other_plan(plan_id):
    if 'user' not in session:
        return redirect('/')
    
    details = OtherPlanDetail.query.filter_by(plan_id=plan_id).all()
    
    data = []
    for detail in details:
        content = json.loads(detail.content)
        content.update({
            'id': detail.id,
            'status': detail.status,
            'operators': ','.join(detail.operators.split(',')),
            'deadline': detail.deadline.strftime('%Y-%m-%d') if detail.deadline else '',
            'completion_time': detail.completion_time.strftime('%Y-%m-%d') if detail.completion_time else ''
        })
        data.append(content)
    
    df = pd.DataFrame(data)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='计划详情')
        worksheet = writer.sheets['计划详情']
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
    
    output.seek(0)
    filename = f"other_plan_{plan_id}_details.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name=filename,
        as_attachment=True
    )

@app.route('/delete_other_plan/<int:plan_id>')
def delete_other_plan(plan_id):
    if 'user' not in session or not session.get('is_admin'):
        return redirect('/')
    
    try:
        # 删除关联的明细条目
        OtherPlanDetail.query.filter_by(plan_id=plan_id).delete()
        # 删除主计划
        OtherPlan.query.filter_by(id=plan_id).delete()
        db.session.commit()
        return redirect('/plan_management')
    except Exception as e:
        db.session.rollback()
        print(f"删除失败: {str(e)}")
        return f"删除失败: {str(e)}", 500

@app.route('/add_other_plan_row/<int:plan_id>', methods=['POST'])
def add_other_plan_row(plan_id):
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({'status': 'error', 'message': '未授权'}), 401
    
    try:
        new_detail = OtherPlanDetail(
            plan_id=plan_id,
            content=json.dumps({"任务名称": "新任务"}),
            status='进行中'
        )
        db.session.add(new_detail)
        db.session.commit()
        return jsonify({
            'id': new_detail.id,
            'content': json.loads(new_detail.content),
            'status': new_detail.status,
            'operators': []
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/update_other_plan/<int:detail_id>', methods=['POST'])
def update_other_plan(detail_id):
    if 'user' not in session:
        return jsonify({'status': 'error', 'message': '未登录'}), 401
    
    current_user = User.query.filter_by(username=session['user']).first()
    if not current_user:
        return jsonify({'status': 'error', 'message': '用户不存在'}), 401
    
    detail = OtherPlanDetail.query.get_or_404(detail_id)
    operators = detail.operators.split(',') if detail.operators else []
    
    # 权限验证：管理员或操作人员
    if not (current_user.is_admin or current_user.username in operators):
        return jsonify({'status': 'error', 'message': '未授权操作'}), 403
    
    try:
        update_data = request.get_json()
        original_data = json.loads(detail.content) if detail.content else {}
        
        # 合并新旧数据（保留未修改字段）
        merged_data = {**original_data, **update_data.get('data', {})}
        
        # 处理日期字段（允许空值）
        deadline = None
        completion_time = None
        if '完成时限' in merged_data:
            try:
                deadline_str = merged_data['完成时限']
                deadline = datetime.strptime(deadline_str, '%Y-%m-%d').date() if deadline_str else None
            except (ValueError, TypeError):
                deadline = None
        if '完成时间' in merged_data:
            try:
                completion_str = merged_data['完成时间']
                completion_time = datetime.strptime(completion_str, '%Y-%m-%d').date() if completion_str else None
            except (ValueError, TypeError):
                completion_time = None
        
        # 更新动态字段内容
        detail.content = json.dumps(merged_data, ensure_ascii=False)
        
        # 更新操作人员（仅管理员可修改）
        if current_user.is_admin and 'operators' in update_data:
            detail.operators = ','.join(update_data['operators'])
        
        # 更新日期字段（允许设置为空）
        detail.deadline = deadline
        detail.completion_time = completion_time
        
        # 自动计算状态
        current_date = datetime.now().date()
        if detail.completion_time:
            if detail.deadline:
                detail.status = '逾期完成' if detail.completion_time > detail.deadline else '已完成'
            else:
                detail.status = '已完成'
        else:
            if detail.deadline:
                detail.status = '逾期' if detail.deadline < current_date else '进行中'
            else:
                detail.status = '进行中'
        
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'data': {
                'id': detail.id,
                'content': json.loads(detail.content),
                'status': detail.status,
                'operators': detail.operators.split(',') if detail.operators else [],
                'deadline': detail.deadline.isoformat() if detail.deadline else None,
                'completion_time': detail.completion_time.isoformat() if detail.completion_time else None
            }
        })
    
    except SQLAlchemyError as e:
        db.session.rollback()
        app.logger.error(f"数据库更新失败: {str(e)}")
        return jsonify({'status': 'error', 'message': '数据库操作失败'}), 500
    except json.JSONDecodeError as e:
        app.logger.error(f"JSON解析失败: {str(e)}")
        return jsonify({'status': 'error', 'message': '数据格式错误'}), 400
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"系统异常: {str(e)}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/delete_other_plan_row/<int:detail_id>', methods=['DELETE'])
def delete_other_plan_row(detail_id):
    if 'user' not in session or not session.get('is_admin'):
        return jsonify({'status': 'error', 'message': '未授权'}), 401
    
    try:
        detail = db.session.get(OtherPlanDetail, detail_id)
        if detail:
            plan_id = detail.plan_id
            db.session.delete(detail)
            db.session.commit()
            return jsonify({
                'status': 'success',
                'redirect': f'/other_plan_detail/{plan_id}'
            })
        return jsonify({'status': 'error', 'message': '条目不存在'}), 404
    
    except SQLAlchemyError as e:
        db.session.rollback()
        app.logger.error(f"数据库删除错误: {str(e)}")
        return jsonify({'status': 'error', 'message': '数据库操作失败'}), 500
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"其他计划条目删除错误: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/weekly_management')
def weekly_management():
    if 'user' not in session:
        return redirect('/')
    return render_template('weekly_management.html')

@app.route('/historical_weekly')
def historical_weekly():
    if 'user' not in session:
        return redirect('/')
    
    current_user = User.query.filter_by(username=session['user']).first()
    page = request.args.get('page', 1, type=int)
    pagination = HistoricalWeeklyReport.query.order_by(
        HistoricalWeeklyReport.page.desc()
    ).paginate(page=page, per_page=1)
    
    reports = []
    archive_time = None 
    if pagination.items:
        content = json.loads(pagination.items[0].content)
        # 获取归档时间
        archive_time = datetime.fromisoformat(content['archive_time']).strftime('%Y-%m-%d %H:%M')
        departments = sorted(set(r['department'] for r in content['reports']))
        for dept in departments:
            dept_reports = [r for r in content['reports'] if r['department'] == dept]
            users = User.query.filter(User.username.in_([r['username'] for r in dept_reports])).all()
            sorted_users = sorted(users, key=lambda u: not u.is_department_head)
            reports.extend([
                {
                    'department': u.department,
                    'username': u.username,
                    'current_work': next(r for r in dept_reports if r['username'] == u.username)['current_work'],
                    'next_plan': next(r for r in dept_reports if r['username'] == u.username)['next_plan'],
                    'submit_time': next(r for r in dept_reports if r['username'] == u.username).get('submit_time')
                }
                for u in sorted_users
            ])
    
    return render_template('historical_weekly.html',
                         reports=reports,
                         pagination=pagination,
                         current_user=current_user,
                         archive_time=archive_time) 

@app.route('/delete_historical_weekly/<int:page_id>', methods=['POST'])
def delete_historical_weekly(page_id):
    if 'user' not in session or not session.get('is_admin'):
        return redirect('/')
    
    report = HistoricalWeeklyReport.query.filter_by(page=page_id).first()
    if report:
        db.session.delete(report)
        db.session.commit()
        flash('历史周报删除成功', 'success')
    else:
        flash('周报记录不存在', 'error')
    
    return redirect(url_for('historical_weekly'))


@app.route('/personal_weekly', methods=['GET', 'POST'])
def personal_weekly():
    if 'user' not in session:
        return redirect('/')
    
    user = User.query.filter_by(username=session['user']).first()
    template = PersonalWeeklyTemplate.query.first()
    
    users = User.query.filter_by(is_company_leader=False).order_by(
        User.department,
        User.is_department_head.desc(),
        User.username
    ).all()
    
    current_reports = {r.user_id: r for r in PersonalWeeklyReport.query.filter_by(archived=False)}
    
    if request.method == 'POST':
        if user.is_company_info or user.is_general_dept_head:
            current_work = request.form.get('current_work', '')
            next_plan = request.form.get('next_plan', '')
            
            if template is None:
                template = PersonalWeeklyTemplate(
                    current_work=current_work,
                    next_plan=next_plan
                )
                db.session.add(template)
            else:
                template.current_work = current_work
                template.next_plan = next_plan
            db.session.commit()
            flash('模板保存成功', 'success')
        
        for u in users:
            if f'submit_{u.id}' in request.form:
                report = current_reports.get(u.id)
                
                if not report:
                    report = PersonalWeeklyReport(
                        user_id=u.id,
                        current_work=request.form.get(f'current_work_{u.id}', ''),
                        next_plan=request.form.get(f'next_plan_{u.id}', '')
                    )
                    db.session.add(report)
                else:

                    if check_permission(user, u) or u.id == user.id:
                        report.current_work = request.form.get(f'current_work_{u.id}', '')
                        report.next_plan = request.form.get(f'next_plan_{u.id}', '')

                if u.id == user.id and not report.submitted: 
                    report.submitted = True
                    report.submit_time = datetime.utcnow()
                elif (user.is_department_head or user.is_department_info) and user.department == u.department:
                    report.department_submitted = True
                    report.submit_time = datetime.utcnow() 
                
                try:
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    flash(f'提交失败: {str(e)}', 'error')
        
        return redirect(url_for('personal_weekly'))
    
    show_archive_button = (user.is_company_info or user.is_general_dept_head)
    return render_template('personal_weekly.html',
                        users=users,
                        current_reports=current_reports,
                        template=template,
                        user=user,
                        show_archive_button=show_archive_button)

@app.route('/archive_weekly', methods=['POST'])
def archive_weekly():
    if 'user' not in session:
        return redirect('/')
    
    user = User.query.filter_by(username=session['user']).first()
    if not (user.is_company_info or user.is_general_dept_head):
        return "无操作权限", 403
    
    current_reports = PersonalWeeklyReport.query.filter_by(archived=False).all()
    for report in current_reports:
        report.archived = True
        report.report_date = datetime.utcnow().date()
    
    history = HistoricalWeeklyReport(
        content=json.dumps({
            'reports': [
                {
                    'department': report.user.department,  # 修复这里，通过关系访问
                    'username': report.user.username,    # 修复这里
                    'current_work': report.current_work,
                    'next_plan': report.next_plan,
                    'submit_time': report.submit_time.isoformat() if report.submit_time else None
                }
                for report in current_reports
            ],
            'archive_time': datetime.utcnow().isoformat()
        }),
        page=HistoricalWeeklyReport.query.count() + 1
    )
    
    db.session.add(history)
    db.session.commit()
    
    return redirect('/historical_weekly')

@app.route('/export_weekly')
def export_weekly():
    if 'user' not in session:
        return redirect('/')
    
    current_date = datetime.now().date()
    year, week_number, _ = current_date.isocalendar()
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"个人周报_{year}年第{week_number}周_{timestamp}.xlsx"
    
    # 获取所有用户（排除公司领导），并按照部门排序
    users = User.query.filter_by(is_company_leader=False).order_by(
        User.department,
        User.is_department_head.desc(),
        User.username
    ).all()
    
    data = []
    for user in users:
        # 查询用户最新的未归档周报
        report = PersonalWeeklyReport.query.filter_by(
            user_id=user.id,
            archived=False
        ).first()
        
        data.append({
            '部门': user.department,
            '姓名': user.username,
            '本周工作内容': report.current_work if report else '',
            '下周计划': report.next_plan if report else '',
            '提交状态': '已提交' if report and report.submitted else '未提交',
            '提交时间': report.submit_time.strftime('%Y-%m-%d %H:%M') if report and report.submit_time else ''
        })
    
    df = pd.DataFrame(data)
    
    # 使用正确的Excel生成方式
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='个人周报')
        worksheet = writer.sheets['个人周报']
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 50
        worksheet.column_dimensions['D'].width = 50
    
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name=filename,
        as_attachment=True
    )

@app.route('/company_weekly', methods=['GET', 'POST'])
def company_weekly():
    if 'user' not in session:
        return redirect('/')
    
    user = User.query.filter_by(username=session['user']).first()
    departments = db.session.query(User.department).filter(
        User.department != '领导'  # 添加过滤条件
    ).distinct().all()
    
    # 获取当前未归档的周报
    current_reports = CompanyWeeklyReport.query.filter_by(archived=False).all()
    
    if request.method == 'POST':
        department = request.form['department']
        report = CompanyWeeklyReport.query.filter_by(department=department, archived=False).first()
        
        if not report:
            report = CompanyWeeklyReport(department=department)
            db.session.add(report)
        
        # 权限检查
        if check_company_write_permission(user, department):
            report.current_work = request.form.get('current_work', '')
            report.next_plan = request.form.get('next_plan', '')
            
            # 如果是部门用户首次提交
            if user.department == department and not report.submitted:
                report.submitted = True
                report.submitter = user.username
                report.submit_time = datetime.utcnow()
            
            db.session.commit()
        
        return redirect(url_for('company_weekly'))
    
    return render_template('company_weekly.html',
                         current_reports=current_reports,
                         user=user,
                         departments=[d[0] for d in departments])

def check_company_write_permission(user, department):
    # 公司信息员和综合部负责人可随时修改
    if user.is_company_info or user.is_general_dept_head:
        return True
    # 部门用户只能修改本部门未提交的
    return user.department == department and not CompanyWeeklyReport.query.filter_by(
        department=department, archived=False, submitted=True).first()

@app.route('/archive_company_weekly', methods=['POST'])
def archive_company_weekly():
    if 'user' not in session:
        return redirect('/')
    
    user = User.query.filter_by(username=session['user']).first()
    if not (user.is_company_info or user.is_general_dept_head):
        return "无操作权限", 403
    
    reports = CompanyWeeklyReport.query.filter_by(archived=False, submitted=True).all()
    for report in reports:
        report.archived = True
        report.archive_time = datetime.utcnow()  # 修正为正确的字段名称
    
    db.session.commit()
    return redirect(url_for('company_weekly'))

@app.route('/company_weekly_history')
def company_weekly_history():
    if 'user' not in session:
        return redirect('/')
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    # 获取所有归档时间点（按归档时间分组）
    archive_query = db.session.query(
        CompanyWeeklyReport.archive_time,
        func.max(CompanyWeeklyReport.id).label('latest_id')
    ).filter(
        CompanyWeeklyReport.archived == True
    ).group_by(
        CompanyWeeklyReport.archive_time
    ).order_by(
        CompanyWeeklyReport.archive_time.desc()
    )
    
    archives = archive_query.paginate(page=page, per_page=per_page, error_out=False)
    
    # 组织数据用于模板展示
    archive_data = []
    for item in archives.items:
        # 获取该归档时间的所有周报
        reports = CompanyWeeklyReport.query.filter(
            CompanyWeeklyReport.archive_time == item.archive_time
        ).order_by(CompanyWeeklyReport.department).all()
        
        archive_data.append({
            'archive_time': item.archive_time,
            'latest_id': item.latest_id,  # 用于删除操作的ID参数
            'reports': reports
        })
    
    return render_template('company_weekly_history.html', 
                         archives=archives,
                         archive_data=archive_data,
                         is_admin=session.get('is_admin', False))

@app.route('/delete_company_weekly_archive/<string:archive_time>')
def delete_company_weekly_archive(archive_time):
    if 'user' not in session or not session.get('is_admin'):
        return redirect('/')
    
    try:
        # 将URL参数中的字符串转换为datetime对象
        archive_time = datetime.fromisoformat(archive_time)
        # 删除所有该归档时间的记录
        CompanyWeeklyReport.query.filter_by(archive_time=archive_time).delete()
        db.session.commit()
        flash('历史记录删除成功', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败: {str(e)}', 'error')
    
    return redirect(url_for('company_weekly_history'))

@app.route('/export_company_weekly_archive/<string:archive_time>')
def export_company_weekly_archive(archive_time):
    if 'user' not in session:
        return redirect('/')
    
    try:
        archive_time = datetime.fromisoformat(archive_time)
        reports = CompanyWeeklyReport.query.filter_by(archive_time=archive_time).all()
        
        data = [{
            '部门': r.department,
            '本周工作': r.current_work,
            '下周计划': r.next_plan,
            '提交时间': r.submit_time.strftime('%Y-%m-%d %H:%M'),
            '归档时间': r.archive_time.strftime('%Y-%m-%d %H:%M')
        } for r in reports]

        df = pd.DataFrame(data)
        filename = f"company_weekly_{archive_time.strftime('%Y%m%d_%H%M')}.xlsx"
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='公司周报')
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            download_name=filename,
            as_attachment=True
        )
    except Exception as e:
        flash(f'导出失败: {str(e)}', 'error')
        return redirect(url_for('company_weekly_history'))

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

with app.app_context():
    db.create_all()
    db.session.execute(text('PRAGMA foreign_keys = ON'))
    db.session.commit()
    
    if not User.query.filter_by(username='admin').first():
        admin = User(
            username='admin',
            password='admin23gg',
            is_admin=True,
            department='测试',
            is_company_info=True,
            is_general_dept_head=True
        )
        db.session.add(admin)
    
    if not OtherPlan.query.first():
        default_plan = OtherPlan(plan_name='默认计划')
        db.session.add(default_plan)
    
    db.session.commit()

if __name__ == '__main__':
    port = int(os.environ.get('FLASK_PORT', 5000))
    host = os.environ.get('FLASK_HOST', '0.0.0.0')
    app.run(host=host, port=port, debug=False)